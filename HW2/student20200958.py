#!/usr/bin/python3

import openpyxl
import math

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total = []
row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append([row_id, sum_v])
	row_id += 1

#grade구하기

total.sort(key=lambda x:x[1], reverse=True)
num = len(total)

a0 = math.floor(num * 0.3)
ap = math.floor(a0 * 0.5)
b0 = math.floor(num * 0.7)
bp = a0 + math.floor((b0 - a0) * 0.5)
cp = math.floor(num * 0.85)

for i in range (num):
        ws.cell(row = total[i][0], column = 8).value = 'C0'
for i in range(int(cp)):
        ws.cell(row = total[i][0], column = 8).value = 'C+'
for i in range(int(b0)):
        ws.cell(row = total[i][0], column = 8).value = 'B0'
for i in range(int(bp)):
        ws.cell(row = total[i][0], column = 8).value = 'B+'
for i in range(int(a0)):
        ws.cell(row = total[i][0], column = 8).value = 'A0'
for i in range(int(ap)):
        ws.cell(row = total[i][0], column = 8).value = 'A+'

wb.save("student.xlsx")
